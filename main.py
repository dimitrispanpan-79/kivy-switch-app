from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.button import Button
from kivy.uix.textinput import TextInput
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner
from kivy.core.window import Window
from kivy.uix.checkbox import CheckBox
from kivy.properties import ListProperty
from kivy.graphics import Color, Rectangle, Line
from kivy.utils import platform

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

import json
import os


Window.clearcolor = (1, 1, 1, 1)

boxes_data = []


class ColoredCheckBox(CheckBox):

    box_color = ListProperty([0.7, 0.85, 1, 1])
    border_color = ListProperty([1, 0, 0, 1])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)

        with self.canvas.before:
            Color(*self.box_color)
            self.rect = Rectangle(pos=self.pos, size=self.size)

            Color(*self.border_color)
            self.line = Line(rectangle=(*self.pos, *self.size), width=1.5)

        self.bind(pos=self.update_rect, size=self.update_rect)

    def update_rect(self, *args):
        self.rect.pos = self.pos
        self.rect.size = self.size
        self.line.rectangle = (*self.pos, *self.size)


def popup(title, text):

    p = Popup(title=title, size_hint=(0.7, 0.35))

    layout = BoxLayout(orientation="vertical", padding=10, spacing=10)

    lbl = Label(
        text=text,
        color=(1, 0, 0, 1),
        font_size="20sp"
    )

    btn = Button(
        text="OK",
        size_hint_y=None,
        height=60,
        background_normal="",
        background_color=(1, 1, 0, 1),
        color=(0, 0, 0, 1),
        font_size="20sp",
        bold=True
    )

    btn.bind(on_press=p.dismiss)

    layout.add_widget(lbl)
    layout.add_widget(btn)

    p.add_widget(layout)
    p.open()


def save_json():

    if platform == "android":
        path = "/storage/emulated/0/Download/switches_data.json"
    else:
        path = "switches_data.json"

    with open(path, "w", encoding="utf-8") as f:
        json.dump(boxes_data, f, ensure_ascii=False, indent=2)

    popup("Saved", "Data saved successfully")


def open_json():

    if platform == "android":
        path = "/storage/emulated/0/Download/switches_data.json"
    else:
        path = "switches_data.json"

    if not os.path.exists(path):
        popup("Error", "File not found")
        return

    global boxes_data

    with open(path, "r", encoding="utf-8") as f:
        boxes_data = json.load(f)

    popup("Loaded", "Data loaded")


def excel_export():

    wb = Workbook()
    ws = wb.active
    ws.title = "Switches"

    headers = ["ROOM", "BOX", "QTY"] + [f"S{i}" for i in range(1, 11)]
    ws.append(headers)

    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, len(headers) + 1):

        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.border = border
        ws.column_dimensions[c.column_letter].width = 15

    for box in boxes_data:

        row = [box["room"], box["box"], box["quantity"]]

        switches = box["switches"][:]

        while len(switches) < 10:
            switches.append("")

        row.extend(switches)
        ws.append(row)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    if platform == "android":
        path = "/storage/emulated/0/Download/Switches.xlsx"
    else:
        path = "Switches.xlsx"

    wb.save(path)

    popup("Saved", f"Excel saved to {path}")


class SwitchApp(App):

    def build(self):

        self.box_number = 1
        self.quantity = 5
        self.switch_entries = []
        self.switch_checks = []

        root = BoxLayout(orientation="horizontal")

        sidebar = BoxLayout(
            orientation="vertical",
            size_hint=(0.3, 1),
            spacing=10,
            padding=5
        )

        buttons = [

            ("NEXT BOX", self.next_box),
            ("FINISH ROOM", self.finish_room),
            ("UNDO LAST", self.undo_last),
            ("PREVIEW", self.preview_list),
            ("SAVE", lambda x: save_json()),
            ("OPEN", lambda x: open_json()),
            ("EXPORT", lambda x: excel_export())

        ]

        for text, func in buttons:

            b = Button(
                text=text,
                background_normal="",
                background_color=(1, 1, 0, 1),
                color=(0, 0, 0, 1),
                font_size="20sp",
                bold=True
            )

            b.bind(on_press=func)
            sidebar.add_widget(b)

        root.add_widget(sidebar)

        main_scroll = ScrollView(size_hint=(0.7, 1))

        self.main = GridLayout(
            cols=1,
            spacing=10,
            padding=10,
            size_hint_y=None
        )

        self.main.bind(minimum_height=self.main.setter("height"))

        self.room_input = TextInput(
            hint_text="ΔΩΜΑΤΙΟ",
            multiline=False,
            size_hint_y=None,
            height=50
        )

        self.main.add_widget(self.room_input)

        self.box_label = Label(
            text="ΝΟ. ΚΟΥΤΙΟΥ: 1",
            color=(0, 0, 0, 1),
            font_size="20sp",
            size_hint_y=None,
            height=40
        )

        self.main.add_widget(self.box_label)

        self.spinner = Spinner(
            text="5",
            values=[str(i) for i in range(1, 11)],
            size_hint_y=None,
            height=50
        )

        self.spinner.bind(text=self.quantity_changed)

        self.main.add_widget(Label(text="ΠΟΣΟΤΗΤΑ", color=(0, 0, 0, 1)))
        self.main.add_widget(self.spinner)

        self.switch_layout = GridLayout(
            cols=1,
            spacing=8,
            size_hint_y=None
        )

        self.switch_layout.bind(minimum_height=self.switch_layout.setter("height"))

        self.main.add_widget(self.switch_layout)

        main_scroll.add_widget(self.main)
        root.add_widget(main_scroll)

        self.create_switches()

        return root

    def quantity_changed(self, spinner, value):
        self.quantity = int(value)
        self.create_switches()

    def create_switches(self):

        self.switch_layout.clear_widgets()

        self.switch_entries = []
        self.switch_checks = []

        for i in range(self.quantity):

            row = BoxLayout(size_hint_y=None, height=50, spacing=5)

            lbl = Label(text=f"Switch {i+1}", color=(0, 0, 0, 1), size_hint=(0.2, 1))
            entry = TextInput(multiline=False)

            chk = ColoredCheckBox(size_hint=(None, None), size=(40, 40))
            chk.bind(active=lambda cb, val, idx=i: self.two_way(idx))

            chk_lbl = Label(text="2 ΘΕΣΕΙΣ", color=(0, 0, 0, 1), size_hint=(0.3, 1))

            row.add_widget(lbl)
            row.add_widget(entry)
            row.add_widget(chk)
            row.add_widget(chk_lbl)

            self.switch_entries.append(entry)
            self.switch_checks.append(chk)

            self.switch_layout.add_widget(row)

    def two_way(self, idx):

        nxt = idx + 1

        if nxt < len(self.switch_entries):
            self.switch_entries[nxt].disabled = self.switch_checks[idx].active

    def save_box(self):

        room = self.room_input.text.strip()

        if room == "":
            popup("Warning", "Write room name")
            return

        switches = []
        i = 0

        while i < len(self.switch_entries):

            txt = self.switch_entries[i].text

            if self.switch_checks[i].active:
                switches.append(txt + " 2Θ")
                switches.append("")
                i += 2
            else:
                switches.append(txt)
                i += 1

        switches = switches[:self.quantity]

        boxes_data.append({
            "room": room,
            "box": self.box_number,
            "quantity": self.quantity,
            "switches": switches
        })

        self.box_number += 1
        self.box_label.text = f"ΝΟ. ΚΟΥΤΙΟΥ: {self.box_number}"

    def next_box(self, instance):
        self.save_box()
        self.clear_fields()

    def finish_room(self, instance):

        self.save_box()

        self.room_input.text = ""
        self.box_number = 1
        self.box_label.text = "ΝΟ. ΚΟΥΤΙΟΥ: 1"

        self.clear_fields()

    def undo_last(self, instance):

        if boxes_data:
            boxes_data.pop()
            self.box_number -= 1

            if self.box_number < 1:
                self.box_number = 1

            self.box_label.text = f"ΝΟ. ΚΟΥΤΙΟΥ: {self.box_number}"

    def clear_fields(self):

        for e in self.switch_entries:
            e.text = ""
            e.disabled = False

        for c in self.switch_checks:
            c.active = False

    def preview_list(self, instance):

        p = Popup(title="Preview", size_hint=(0.9, 0.9))

        scroll = ScrollView()

        layout = GridLayout(cols=1, spacing=5, size_hint_y=None)
        layout.bind(minimum_height=layout.setter("height"))

        for b in boxes_data:

            line = f"{b['room']} | Box {b['box']} | {', '.join([s for s in b['switches'] if s])}"

            layout.add_widget(Label(
                text=line,
                size_hint_y=None,
                height=40,
                color=(0, 0, 0, 1)
            ))

        scroll.add_widget(layout)
        p.add_widget(scroll)
        p.open()


if __name__ == "__main__":
    SwitchApp().run()